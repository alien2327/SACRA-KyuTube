from openpyxl import load_workbook
import glob
import pickle

from pages import *

header = {
    'A':'index', 'B':'classroom', 'C':'subject', 'D':'is_video',
    'E':'data_type', 'F':'title_jp', 'G':'title_en', 'H':'bio[0].phylum',
    'I':'bio[0].bio_class', 'J':'bio[0].name_jp', 'K':'bio[0].name_en', 
    'P':'bio_type', 'Q':'capture_date', 'R':'capture_time', 'S':'capture_place.place_type',
    'T':'capture_place.is_local', 'U':'capture_place.location_1', 
    'V':'capture_place.location_2', 'W':'capture_place.location_1', 
    'X':'capture_place.location_2', 'AB':'condition.use_micro', 
    'AC':'condition.use_auto', 'AD':'condition.use_probe', 
    'AE':'condition.use_infrared', 'AF':'condition.capture_interval', 
    'AG':'condition.capture_speed', 'AH':'condition.photographer', 
    'AI':'condition.copyright', 'AJ':'is_opened', 'AK':'comment_jp',
    'AL':'comment_en'
}

def read_pkl():
    file = glob.glob('pkl/*.pkl')
    data = [pickle.load(open(f, 'rb')) for f in file]
    return data

def update(fname):
    print('Start updating excel file')
    wb = load_workbook('./template.xlsx')
    sheet = wb[wb.sheetnames[0]]
    data = read_pkl()
    for i, d in enumerate(data):
        i = i + 3
        for k, v in header.items():
            if k == 'D':
                if eval(f'd.{v}'): sheet[k+str(i)] = '静止画'
                else: sheet[k+str(i)] = '動画'
            elif k == 'AJ': sheet[k+str(i)] = '学内外'
            else:
                if eval(f'd.capture_place.is_local') == '国内':
                    if k == 'W' or k == 'X': continue
                    else: sheet[k+str(i)] = eval(f'd.{v}')
                elif eval(f'd.capture_place.is_local') == '国外':
                    if k == 'U' or k == 'V': continue
                    else: sheet[k+str(i)] = eval(f'd.{v}')
                else: sheet[k+str(i)] = eval(f'd.{v}')
    wb.save(fname)
    print('End updating excel file')

if __name__ == "__main__":
    update('./kyutube.xlsx')